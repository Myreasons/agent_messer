import pandas as pd
import psycopg2
import datetime as dt
import os
from nptime import nptime
import sys
sys.path.insert(0, r'C:\Users\Admin\Desktop\Projects\agent messer\sub lib')
import openpyxl
from openpyxl.styles import PatternFill

from outlook_manager import Mail


def get_jf(period_start, period_end):
	con = psycopg2.connect(
	  database="sue", 
	  user="read_user", 
	  password="pDs7Qj21", 
	  host="172.19.16.130", 
	  port="5432"
	)

	query = '''SELECT
	mrf,
	comment,
	event_end,
	id,
	influence,
	event_start,
	service,
	opisanie,
	"Downtome_employee_count",
	"GP appeals count",
	"worst_sl",
	"worst_lcr",
	"meropriyatia"
	FROM public."48form_for_bi_all_data"
	WHERE event_start BETWEEN '%s' ''' % period_start + '''AND '%s' ''' % period_end +'''
	ORDER by event_start DESC'''
	

	df_jf = pd.read_sql(query, con)
	con.close()

	df_jf['event_end'] = pd.to_datetime(df_jf['event_end'])
	df_jf['event_start'] = pd.to_datetime(df_jf['event_start'])
	df_jf['Дата'] = pd.to_datetime(df_jf['event_start'])
	df_jf['Дата'] = df_jf['Дата'].apply(lambda x: x.date())
	df_jf['Время простоя'] = 0


	for ind in df_jf.index.values:
		df_jf['Время простоя'].iloc[ind] = (df_jf['event_end'].iloc[ind] - df_jf['event_start'].iloc[ind]).total_seconds()

	df_jf['Время простоя'] = df_jf['Время простоя']/60

	df_jf['Время зависания'] = df_jf['Время простоя']
	df_jf['Время простоя'] = df_jf['Время простоя']*df_jf['Downtome_employee_count']
	
	df_jf.rename(columns={'mrf': 'МРФ', 'comment': 'Описание', 
					'event_start': 'Дата и время  начала проблемы', 'event_end': 'Дата и время  окончания проблемы', 
					'GP appeals count': 'Количество обращений', 'Downtome_employee_count': 'Количество затронутых операторов', 
					'id': 'Номер в JF', '': 'Статус проблемы', 
					'worst_sl': 'SL - Наихудший показатель в течение события/аварии', 
					'worst_lcr': 'LCR - Наихудший показатель в течение события/аварии', 'meropriyatia': 'Корректирующие мероприятия',
					'opisanie': 'Доп комментарий', 'influence': 'Статус проблемы'}, inplace=True)
	
	df_jf['Номер недели'] = df_jf['Дата'].apply(lambda x: x.isocalendar()[1])
	df_jf = df_jf[['Дата', 'Номер недели', 'МРФ', 'Описание', 'Дата и время  начала проблемы', 
	'Дата и время  окончания проблемы', 'Время простоя', 'Время зависания', 
	'Количество обращений', 'Количество затронутых операторов', 'Номер в JF','Статус проблемы', 
	'SL - Наихудший показатель в течение события/аварии', 'LCR - Наихудший показатель в течение события/аварии', 
	'Доп комментарий', 'Корректирующие мероприятия']]
	df_jf = df_jf.set_index(['Дата'], drop = True)

	return(df_jf)

"""
Дата
МРФ
Аварий на сетях МРФ
Дата и время  начала проблемы
Дата и время  окончания проблемы
Время простоя
Время зависания
Количество обращений/Количество затронутых операторов
Номер в JF
Статус проблемы
SL - Наихудший показатель в течение события/аварии
LCR - Наихудший показатель в течение события/аварии
Доп комментарий
Чья зона ответственности
Корректирующие мероприятия
"""

def create_jf_report():
	end = dt.date.today() - dt.timedelta(days = 1)
	start = end - dt.timedelta(days = 7)

	df = get_jf(start, end)

	writer = pd.ExcelWriter('gp_report.xlsx', engine='xlsxwriter')
	os.chdir(r'C:\Users\Admin\Desktop\Projects\agent messer\sent_mail')
	df.to_excel(writer, 'report')
	writer.save()

def as_text(value):
    if value is None:
        return ""
    return str(value)

"""
Высокое
Максимальное
Низкое
Среднее
"""

def customize_excel(path):

	style_dict = {
	'Максимальное': PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid'),
	'Среднее': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
	'Высокое': PatternFill(start_color='FF9200', end_color='FF9200', fill_type='solid')
	}

	wb = openpyxl.load_workbook(path)
	ws=wb.active
	for column_cells in ws.columns:
		length = max(len(as_text(cell.value)) for cell in column_cells)
		if length > 33:
			length = 33
		ws.column_dimensions[column_cells[0].column_letter].width = length
	
	for row in range(1, ws.max_row + 1):
		try:
			for col in range(1, ws.max_column+1):
				ws.cell(row = row, column = col).fill = style_dict[ws.cell(row = row, column = 12).value]
		except:
			continue

	wb.save(path)
create_jf_report()
customize_excel(r'C:\Users\Admin\Desktop\Projects\agent messer\sent_mail\gp_report.xlsx')